# ================================================================================
# HEADER
# ================================================================================
"""
File: 3model_response_verifier.py
Author: Dr. Shradha's Research Team
Date: 2025-12-11
Version: 1.1.0
Introduction: This script verifies model responses against ground truth answers
using an LLM agent, adding correctness flags and explanations to the JSON file.
"""

# ================================================================================
# IMPORTS
# ================================================================================
import pandas as pd
import asyncio
import json
import logging
from typing import Dict, Any, Optional
from google.adk.agents import LlmAgent
from google.adk.runners import Runner
from google.adk.sessions import InMemorySessionService
from google.genai import types
from pydantic import BaseModel, Field
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from models import get_model_for_agent
from textwrap import dedent

os.chdir(os.path.dirname(os.path.abspath(__file__)))

MODEL = get_model_for_agent()

# ================================================================================
# FILE DEFINITIONS
# ================================================================================
RESPONSES_JSON_FILE = "data/qna_responses_final.json"
VERIFIED_JSON_FILE = "data/qna_verified_responses.json"
OUTPUT_EXCEL_FILE = "data/qna_verified_responses_final.xlsx"
OUTPUT_EXCEL_WITH_EXPLAIN = "data/qna_verified_responses_final_with_explain.xlsx"

# ================================================================================
# CONFIGURATION
# ================================================================================
def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    import logging as log_module
    litellm_logger = log_module.getLogger("LiteLLM")
    litellm_logger.setLevel(log_module.WARNING)

    return logging.getLogger(__name__)


class VerificationResult(BaseModel):
    """Pydantic model for verification response validation"""
    is_correct: bool = Field(description="True if response matches correct answer")
    explanation: str = Field(description="Brief explanation of the decision")


agent_runner = None
session_id = None


def initialize_agent_runner():
    """Initialize the agent runner (sessions will be created per question)"""
    global agent_runner
    if agent_runner is None:
        session_service = InMemorySessionService()
        verification_agent = create_verification_agent(MODEL)
        agent_runner = Runner(
            agent=verification_agent,
            app_name="agents",
            session_service=session_service
        )


def create_verification_agent(model):
    """Create a verification agent that outputs JSON via instructions"""
    # return LlmAgent(
    #     model=model,
    #     name="verification_agent",
    #     description="Verifies if a model response for a linear-agebra problem matches the correct answer.",
    #     instruction=dedent("""
    #         You are an expert in linear algebra / mathematics.
    #         Yoor task is to correctly verify if the model response for given linear algebra / mathematics problem matches with the given correct answer (latex form).
    #         You will provided your response in JSON format like: {"model_response": "...", "correct_answer": "..."}.
    #         While comparing, consider equivalent mathematical expressions, simplified forms, or different representations that are conceptually identical as correct.

    #         IMPORTANT: Respond ONLY with a valid JSON object in this exact format (no markdown, no extra text):
    #         {"is_correct": true, "explanation": "brief explanation"}
    #         or
    #         {"is_correct": false, "explanation": "brief explanation"}
    #     """).strip()
    # )
    return LlmAgent(
            model=model,
            name="verification_agent",
            description="Verifies if a model response conceptually matches a correct mathematical answer.",
            instruction=dedent("""
                You are a precise verification agent that checks if a model response conceptually matches a correct mathematical answer.
                The user will provide a model response and the ground truth answer in JSON format like: {"model_response": "...", "correct_answer": "..."}.
                Compare the model_response to the correct_answer mathematically. Consider equivalent mathematical expressions, simplified forms, or different representations that are conceptually identical as correct.

                IMPORTANT: Respond ONLY with a valid JSON object in this exact format (no markdown, no extra text):
                {"is_correct": true, "explanation": "brief explanation"}
                or
                {"is_correct": false, "explanation": "brief explanation"}
            """).strip()
        )


# ================================================================================
# VERIFICATION FUNCTIONS
# ================================================================================
async def call_verification_agent(model_response, correct_answer, session_id):
    """Call the verification agent to check if the response matches the correct answer"""
    global agent_runner

    # Truncate long responses, keeping ending portion (where final answer typically is)
    max_length = 20000
    if len(model_response) > max_length:
        normalized_model_response = "[TRUNCATED_BEGINNING] " + model_response[-max_length:]
        print('MODEL RESPONSE TOO LONG.')
    else:
        normalized_model_response = model_response

    query_json = json.dumps({
        "model_response": normalized_model_response,
        "correct_answer": correct_answer
    })

    user_content = types.Content(role='user', parts=[types.Part(text=query_json)])

    # Run agent and collect response
    response_text = None
    async for event in agent_runner.run_async(user_id="verifier", session_id=session_id, new_message=user_content):
        if event.is_final_response() and event.content and event.content.parts:
            response_text = event.content.parts[0].text
            break

    if not response_text:
        return {"is_correct": False, "explanation": "No response from agent"}

    # Clean markdown wrapping if present
    response_text = response_text.strip()
    if response_text.startswith("```json"):
        response_text = response_text[7:]
    if response_text.startswith("```"):
        response_text = response_text[3:]
    if response_text.endswith("```"):
        response_text = response_text[:-3]
    response_text = response_text.strip()

    # Parse JSON
    try:
        result = json.loads(response_text)
        return result
    except json.JSONDecodeError:
        return {"is_correct": False, "explanation": f"Failed to parse JSON: {response_text[:100]}"}


async def response_gt_compare(responses_json, question_id, model_name, ground_truth, json_file, session_id):
    """
    Compare model response with ground truth for a specific question and model.
    Skip if verification already exists in the output file.
    Note: Response changes are detected and cleared BEFORE this function is called.
    """
    logger = logging.getLogger(__name__)

    # Get response from new JSON format structure
    response = responses_json[question_id]['responses'].get(model_name)
    if not response or str(response).strip() == '':
        logger.info(f"  {model_name}: No response to verify")
        return None

    existing_result = responses_json[question_id]['correct'].get(model_name)
    if existing_result is not None and isinstance(existing_result, bool):
        logger.info(f"  {model_name}: Verification already exists ({existing_result}), skipping")
        return existing_result

    model_response = responses_json[question_id]['responses'][model_name]
    verification_result = await call_verification_agent(model_response, ground_truth, session_id)

    is_correct = verification_result['is_correct']
    explanation = verification_result['explanation']

    # Store in nested 'correct' structure
    if 'correct' not in responses_json[question_id]:
        responses_json[question_id]['correct'] = {}
    responses_json[question_id]['correct'][model_name] = is_correct

    # Store in nested 'explanations' structure
    if 'explanations' not in responses_json[question_id]:
        responses_json[question_id]['explanations'] = {}
    responses_json[question_id]['explanations'][model_name] = explanation

    logger.info(f"  {model_name}: {'CORRECT' if is_correct else 'INCORRECT'} - {explanation}")

    with open(json_file, 'w') as f:
        json.dump(responses_json, f, indent=2)

    return is_correct

# ================================================================================
# MAIN PROCESSING
# ================================================================================
async def verify_model_responses():
    """
    Verify model responses against the answer_latex using a structured LLM agent.
    Returns True/False for each response match.
    """
    logger = setup_logging()

    logger.info(f"Loading responses from {RESPONSES_JSON_FILE} (READ-ONLY)")
    with open(RESPONSES_JSON_FILE, 'r') as f:
        responses_input = json.load(f)
    logger.info(f"Loaded {len(responses_input)} questions from input")

    verified_exists = os.path.exists(VERIFIED_JSON_FILE)

    if verified_exists:
        logger.info(f"Loading existing VERIFIED_JSON_FILE from {VERIFIED_JSON_FILE}")
        with open(VERIFIED_JSON_FILE, 'r') as f:
            verified_data = json.load(f)
        logger.info(f"Loaded {len(verified_data)} questions from verified cache")
    else:
        logger.info(f"VERIFIED_JSON_FILE not found, creating new from RESPONSES structure")
        verified_data = {}

    logger.info("Syncing VERIFIED structure with RESPONSES structure...")

    # ============================================================================
    # DETECT RESPONSE CHANGES BEFORE SYNCING (Approach 1 - EARLY DETECTION)
    # ============================================================================
    changes_detected = {}

    for question_id, response_data in responses_input.items():
        if 'responses' not in response_data:
            continue

        for model_name, new_response in response_data['responses'].items():
            if new_response is None or str(new_response).strip() == '':
                continue

            # Get old response from verified_data
            old_response = None
            if (question_id in verified_data and
                'responses' in verified_data[question_id] and
                model_name in verified_data[question_id]['responses']):
                old_response = verified_data[question_id]['responses'][model_name]

            # Check if response changed
            if old_response is not None and old_response != new_response:
                if question_id not in changes_detected:
                    changes_detected[question_id] = []
                changes_detected[question_id].append(model_name)
                logger.info(f"  CHANGE DETECTED: {question_id} / {model_name}")

    logger.info(f"Total changes detected: {sum(len(v) for v in changes_detected.values())} model responses")
    # ============================================================================

    for question_id, response_data in responses_input.items():
        if question_id not in verified_data:
            verified_data[question_id] = {}
            logger.info(f"  Added new question: {question_id}")

        # Handle metadata from new structure with metadata key (required)
        if 'metadata' not in verified_data[question_id]:
            verified_data[question_id]['metadata'] = {}
        verified_data[question_id]['metadata'].update(response_data['metadata'])

        if 'responses' not in verified_data[question_id]:
            verified_data[question_id]['responses'] = {}

        if 'responses' in response_data and isinstance(response_data['responses'], dict):
            for model_name, response_text in response_data['responses'].items():
                verified_data[question_id]['responses'][model_name] = response_text

        if 'correct' not in verified_data[question_id]:
            verified_data[question_id]['correct'] = {}

        if 'responses' in verified_data[question_id]:
            for model_name in verified_data[question_id]['responses'].keys():
                if model_name not in verified_data[question_id]['correct']:
                    verified_data[question_id]['correct'][model_name] = None

        # ====================================================================
        # CLEAR VERIFICATIONS FOR CHANGED RESPONSES (Approach 1)
        # ====================================================================
        if question_id in changes_detected:
            for model_name in changes_detected[question_id]:
                if 'correct' in verified_data[question_id]:
                    verified_data[question_id]['correct'][model_name] = None
                if 'explanations' in verified_data[question_id]:
                    verified_data[question_id]['explanations'][model_name] = None
                logger.info(f"  Cleared verification for {question_id} / {model_name}")
        # ====================================================================

    logger.info("Synchronizing manual updates from RESPONSES 'correct' key...")
    synced_updates = 0

    for question_id, response_data in responses_input.items():
        if 'correct' in response_data:
            for model_name, correct_value in response_data['correct'].items():
                if (correct_value is not None and
                    isinstance(correct_value, bool) and
                    verified_data[question_id]['correct'].get(model_name) is None):
                    verified_data[question_id]['correct'][model_name] = correct_value
                    logger.info(f"  Synced manual update: {question_id} / {model_name} = {correct_value}")
                    synced_updates += 1

    logger.info(f"Synced {synced_updates} manual verifications from RESPONSES")

    with open(VERIFIED_JSON_FILE, 'w') as f:
        json.dump(verified_data, f, indent=2)
    logger.info(f"VERIFIED_JSON_FILE initialized and saved")

    verified_count = 0
    for question_id, data in verified_data.items():
        for model_name, value in data['correct'].items():
            if value is not None and isinstance(value, bool):
                verified_count += 1
    logger.info(f"Found {verified_count} existing verified results (agent + manual)")

    logger.info(f"Ground truth (answer_latex) will be read from VERIFIED structure")

    initialize_agent_runner()

    all_models = set()
    for question_id, question_data in verified_data.items():
        if 'responses' in question_data and isinstance(question_data['responses'], dict):
            for model_name, response_value in question_data['responses'].items():
                if (response_value is not None and
                    str(response_value).lower() != 'nan' and
                    str(response_value).strip() != ''):
                    all_models.add(model_name)

    model_names = sorted(list(all_models))
    logger.info(f"Detected {len(model_names)} models to verify: {model_names}")

    questions_in_verified = list(verified_data.keys())
    logger.info(f"Starting verification for {len(questions_in_verified)} questions")

    for index, question_id in enumerate(questions_in_verified):
        ground_truth = verified_data[question_id]['metadata']['answer_latex']
        if not ground_truth:
            logger.warning(f"  Question {question_id} has no answer_latex, skipping")
            continue

        print(f"\\n{'='*80}")
        print(f"Processing Question {index+1} of {len(questions_in_verified)}: {question_id}")
        print(f"{'='*80}")
        logger.info(f"[{index+1}/{len(questions_in_verified)}] Processing question {question_id}")

        # Create fresh session for this question to prevent context accumulation
        question_session_id = f"verification_{question_id}"
        await agent_runner.session_service.create_session(app_name="agents", user_id="verifier", session_id=question_session_id)

        for model_idx, model_name in enumerate(model_names):
            print(f"  Model {model_idx+1}/{len(model_names)}: {model_name}")
            await response_gt_compare(verified_data, question_id, model_name, ground_truth, VERIFIED_JSON_FILE, question_session_id)

        if (index + 1) % 5 == 0:
            print(f"\n{'-'*60}")
            print(f"PROGRESS SUMMARY AFTER {index+1} QUESTIONS")
            print(f"{'-'*60}")
            for model_name in model_names:
                correct_count = 0
                total_count = 0

                for q_id in list(verified_data.keys())[:index+1]:
                    if (isinstance(verified_data[q_id], dict) and
                        'correct' in verified_data[q_id] and
                        isinstance(verified_data[q_id]['correct'], dict) and
                        model_name in verified_data[q_id]['correct']):
                        total_count += 1
                        if verified_data[q_id]['correct'][model_name] is not None:
                            if verified_data[q_id]['correct'][model_name]:
                                correct_count += 1

                if total_count > 0:
                    accuracy = (correct_count / total_count) * 100
                    print(f"  {model_name}: {correct_count}/{total_count} correct ({accuracy:.1f}%)")
                else:
                    print(f"  {model_name}: No verifications yet")

    print(f"\\n{'='*80}")
    print("Verification process completed")
    print(f"{'='*80}")

    with open(VERIFIED_JSON_FILE, 'w') as f:
        json.dump(verified_data, f, indent=2)
    logger.info(f"Verification completed! Final results saved to {VERIFIED_JSON_FILE}")

# ================================================================================
# EXCEL EXPORT
# ================================================================================
    # Extract metadata columns dynamically from the verified data (first question)
    first_question_id = list(verified_data.keys())[0]
    metadata_cols = ['id'] + sorted([col for col in verified_data[first_question_id]['metadata'].keys()])

    df_rows = []
    for question_id, data in verified_data.items():
        row = {}

        # Add question_id as 'id' column (QID is the dictionary key, not a field)
        row['id'] = question_id

        for col in metadata_cols:
            if col != 'id':  # Skip 'id' since we already added it
                row[col] = data['metadata'].get(col, None)

        if 'responses' in data and isinstance(data['responses'], dict):
            for model_name in sorted(data['responses'].keys()):
                row[model_name] = data['responses'].get(model_name, None)
                if 'correct' in data and isinstance(data['correct'], dict):
                    correct_val = data['correct'].get(model_name, None)
                    if correct_val is True:
                        row[f"{model_name}_correct"] = "TRUE"
                    elif correct_val is False:
                        row[f"{model_name}_correct"] = "FALSE"
                    else:
                        row[f"{model_name}_correct"] = ""
                else:
                    row[f"{model_name}_correct"] = ""

        df_rows.append(row)

    result_df = pd.DataFrame(df_rows)
    result_df.to_excel(OUTPUT_EXCEL_FILE, index=False)

    wb = load_workbook(OUTPUT_EXCEL_FILE)
    ws = wb.active

    # Map response columns to their corresponding _correct columns for formatting
    response_to_correct_map = {}
    correct_col_indices = set()

    for idx, col in enumerate(result_df.columns, start=1):
        if col.endswith('_correct'):
            correct_col_indices.add(idx)
            response_col_name = col.replace('_correct', '')
            response_col_idx = list(result_df.columns).index(response_col_name) + 1
            response_to_correct_map[response_col_idx] = idx

    center_top = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    col_max_lengths = {}

    # Single loop: apply alignment, background color, track widths, set heights
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
        ws.row_dimensions[row_idx].height = None if row_idx == 1 else 144

        for col_idx, cell in enumerate(row, start=1):
            cell.alignment = center_top if col_idx in correct_col_indices else left_top

            if row_idx > 1 and col_idx in response_to_correct_map:
                correct_col_idx = response_to_correct_map[col_idx]
                correct_cell = row[correct_col_idx - 1]
                if correct_cell.value == "FALSE":
                    cell.fill = yellow_fill

            if cell.value:
                cell_value = str(cell.value)
                max_line_len = max(len(line) for line in cell_value.split('\n'))
                col_max_lengths[col_idx] = max(col_max_lengths.get(col_idx, 0), max_line_len)

    for col_idx, max_len in col_max_lengths.items():
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = min(max_len + 2, 100)

    wb.save(OUTPUT_EXCEL_FILE)
    logger.info(f"Verification results exported to Excel: {OUTPUT_EXCEL_FILE}")

    print("\\nFinal Verification Summary:")
    for model_name in model_names:
        correct_count = 0
        total_count = 0

        for question_id, data in verified_data.items():
            if (isinstance(data, dict) and
                'correct' in data and
                isinstance(data['correct'], dict) and
                model_name in data['correct']):
                total_count += 1
                if data['correct'][model_name]:
                    correct_count += 1

        if total_count > 0:
            accuracy = (correct_count / total_count) * 100
            print(f"  {model_name}: {correct_count}/{total_count} correct ({accuracy:.1f}%)")

    return verified_data

# ================================================================================
# MAIN EXECUTION
# ================================================================================
async def main():
    """Main function to run the verification"""
    results = await verify_model_responses()
    print(f"\nVerification process completed. Results stored in {VERIFIED_JSON_FILE}")

if __name__ == "__main__":
    asyncio.run(main())
